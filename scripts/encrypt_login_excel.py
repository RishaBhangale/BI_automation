# scripts/encrypt_login_excel.py
"""
One-time script to encrypt the 'password' column in login_testdata.xlsx.

Run this ONCE after implementing encryption_utils.py and setting
TEST_FRAMEWORK_SECRET_KEY in your environment:

    python scripts/encrypt_login_excel.py

After this runs:
- The password column will contain Fernet tokens (encrypted).
- Your tests will still pass because excel_utils.load_login_testdata()
  transparently decrypts on read.
"""
import sys
from pathlib import Path

# Add project root to Python path so 'utils' is importable
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
import shutil
from pathlib import Path
from dotenv import load_dotenv
load_dotenv()

from openpyxl import load_workbook

from utils.encryption_utils import encrypt_value

# Anchor paths to project root regardless of where the script is run from
PROJECT_ROOT = Path(__file__).resolve().parent.parent

EXCEL_PATH   = PROJECT_ROOT / "testdata" / "login_testdata.xlsx"
BACKUP_PATH  = PROJECT_ROOT / "testdata" / "login_testdata_PLAIN_BACKUP.xlsx"
SHEET_NAME = "LoginData"
PASSWORD_COLUMN_NAME = "password"  # must match your Excel header exactly


def main():
    if not EXCEL_PATH.is_file():
        raise FileNotFoundError(f"Excel not found at: {EXCEL_PATH}")

    # --- Step 1: backup the plain file (git-ignore this backup) ---
    shutil.copy2(EXCEL_PATH, BACKUP_PATH)
    print(f"Backup saved to: {BACKUP_PATH}")

    # --- Step 2: find the password column index ---
    wb = load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]

    # Row 1 is the header row
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    headers_lower = [str(h).strip().lower() if h else "" for h in headers]

    if PASSWORD_COLUMN_NAME not in headers_lower:
        raise ValueError(
            f"Column '{PASSWORD_COLUMN_NAME}' not found in sheet '{SHEET_NAME}'. "
            f"Available columns: {headers}"
        )

    # openpyxl columns are 1-based
    pwd_col_index = headers_lower.index(PASSWORD_COLUMN_NAME) + 1

    # --- Step 3: encrypt each password cell (skip header row) ---
    encrypted_count = 0
    skipped_count = 0

    for row in ws.iter_rows(min_row=2):
        cell = row[pwd_col_index - 1]  # iter_rows yields 0-based within the row
        raw = cell.value

        if not raw or str(raw).strip() == "":
            skipped_count += 1
            continue

        # Avoid double-encrypting if already a Fernet token
        raw_str = str(raw).strip()
        if raw_str.startswith("gAAAAA"):
            print(f"  Row {cell.row}: already encrypted, skipping.")
            skipped_count += 1
            continue

        cell.value = encrypt_value(raw_str)
        encrypted_count += 1

    # --- Step 4: save ---
    wb.save(EXCEL_PATH)
    print(f"\nDone. Encrypted: {encrypted_count} rows | Skipped: {skipped_count} rows")
    print(f"Encrypted file saved to: {EXCEL_PATH}")
    print(f"\nIMPORTANT: Add '{BACKUP_PATH}' to .gitignore — it contains plain passwords.")


if __name__ == "__main__":
    main()